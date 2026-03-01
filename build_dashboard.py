"""
Restaurant Operations Dashboard - The Grumpy Chef
Builds a single-workbook Excel dashboard connecting 8 template modules.
"""

import os
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, NamedStyle, numbers
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule

# ── Colour palette ──────────────────────────────────────────────────────────
NAVY       = "0A1628"
DARK_NAVY  = "1B2A4A"
GOLD       = "D4AF37"
WHITE      = "F8F9FA"
RED        = "C53030"
GREEN      = "2F855A"
INPUT_BG   = "FFF2CC"   # yellow tint for input cells
CALC_BG    = "D9D9D9"   # gray for auto-calc cells
BLACK      = "000000"

# ── Reusable style objects ──────────────────────────────────────────────────
thin_border = Border(
    left=Side(style="thin", color=NAVY),
    right=Side(style="thin", color=NAVY),
    top=Side(style="thin", color=NAVY),
    bottom=Side(style="thin", color=NAVY),
)

fill_navy      = PatternFill(start_color=DARK_NAVY, end_color=DARK_NAVY, fill_type="solid")
fill_navy_dark = PatternFill(start_color=NAVY, end_color=NAVY, fill_type="solid")
fill_input     = PatternFill(start_color=INPUT_BG, end_color=INPUT_BG, fill_type="solid")
fill_calc      = PatternFill(start_color=CALC_BG, end_color=CALC_BG, fill_type="solid")
fill_white     = PatternFill(start_color=WHITE, end_color=WHITE, fill_type="solid")
fill_gold      = PatternFill(start_color=GOLD, end_color=GOLD, fill_type="solid")
fill_green     = PatternFill(start_color=GREEN, end_color=GREEN, fill_type="solid")
fill_red       = PatternFill(start_color=RED, end_color=RED, fill_type="solid")

font_title     = Font(name="Calibri", size=20, bold=True, color=GOLD)
font_subtitle  = Font(name="Calibri", size=11, italic=True, color="888888")
font_section   = Font(name="Calibri", size=13, bold=True, color=GOLD)
font_label     = Font(name="Calibri", size=11, bold=True, color=WHITE)
font_value     = Font(name="Calibri", size=11, color=BLACK)
font_value_bold = Font(name="Calibri", size=11, bold=True, color=BLACK)
font_target    = Font(name="Calibri", size=10, italic=True, color="555555")
font_status_g  = Font(name="Calibri", size=11, bold=True, color=GREEN)
font_status_r  = Font(name="Calibri", size=11, bold=True, color=RED)
font_white     = Font(name="Calibri", size=11, color=WHITE)
font_white_b   = Font(name="Calibri", size=11, bold=True, color=WHITE)
font_gold_sm   = Font(name="Calibri", size=10, color=GOLD)

align_left   = Alignment(horizontal="left", vertical="center", wrap_text=True)
align_center = Alignment(horizontal="center", vertical="center")
align_right  = Alignment(horizontal="right", vertical="center")


# ── Helpers ─────────────────────────────────────────────────────────────────
def set_cell(ws, row, col, value, font=None, fill=None, alignment=None,
             border=thin_border, num_fmt=None):
    """Write a value and apply formatting to a single cell."""
    cell = ws.cell(row=row, column=col, value=value)
    if font:
        cell.font = font
    if fill:
        cell.fill = fill
    if alignment:
        cell.alignment = alignment
    else:
        cell.alignment = align_left if col == 1 else align_center
    if border:
        cell.border = border
    if num_fmt:
        cell.number_format = num_fmt
    return cell


def section_header(ws, row, title, cols=4):
    """Gold text on dark navy across columns A-D."""
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=cols)
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = fill_navy
        cell.border = thin_border
    cell_a = ws.cell(row=row, column=1)
    cell_a.value = title
    cell_a.font = font_section
    cell_a.alignment = Alignment(horizontal="left", vertical="center")


def label_row(ws, row, label, formula_or_value=None, target=None, status_formula=None,
              is_input=False, is_pct=False, is_currency=False, is_integer=False):
    """
    Standard KPI row:
      A = label (white on navy)
      B = value (input or calc)
      C = target / benchmark
      D = status
    """
    # Col A - label
    set_cell(ws, row, 1, label, font=font_label, fill=fill_navy_dark)

    # Col B - value
    b_fill = fill_input if is_input else fill_calc
    b_font = font_value_bold if not is_input else font_value
    num_fmt = None
    if is_pct:
        num_fmt = "0.0%"
    elif is_currency:
        num_fmt = '$#,##0'
    elif is_integer:
        num_fmt = '#,##0'
    set_cell(ws, row, 2, formula_or_value, font=b_font, fill=b_fill, num_fmt=num_fmt)

    # Col C - target
    t_fmt = None
    if target is not None and isinstance(target, str) and "%" in str(target):
        t_fmt = None  # keep as string
    set_cell(ws, row, 3, target, font=font_target, fill=fill_white, num_fmt=t_fmt)

    # Col D - status
    if status_formula:
        set_cell(ws, row, 4, status_formula, font=font_value_bold, fill=fill_white)


# ═══════════════════════════════════════════════════════════════════════════
#  SHEET 1 : DASHBOARD
# ═══════════════════════════════════════════════════════════════════════════
def build_dashboard(wb):
    ws = wb.active
    ws.title = "Dashboard"

    # Column widths
    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 18

    # ── Row 1: Title ────────────────────────────────────────────────────
    ws.merge_cells("A1:D1")
    set_cell(ws, 1, 1, "Restaurant Operations Dashboard \u2014 The Grumpy Chef",
             font=font_title, fill=fill_navy_dark,
             alignment=Alignment(horizontal="left", vertical="center"),
             border=Border(bottom=Side(style="medium", color=GOLD)))
    for c in range(2, 5):
        ws.cell(row=1, column=c).fill = fill_navy_dark
        ws.cell(row=1, column=c).border = Border(bottom=Side(style="medium", color=GOLD))
    ws.row_dimensions[1].height = 38

    # ── Row 2: Subtitle ────────────────────────────────────────────────
    ws.merge_cells("A2:D2")
    set_cell(ws, 2, 1, "One Page. One Snapshot. Every Number That Matters.",
             font=font_subtitle, fill=fill_white,
             alignment=Alignment(horizontal="left", vertical="center"))
    for c in range(2, 5):
        ws.cell(row=2, column=c).fill = fill_white
    ws.row_dimensions[2].height = 22

    # ── Row 3: Week-of input ───────────────────────────────────────────
    set_cell(ws, 3, 1, "Week of:", font=font_label, fill=fill_navy_dark)
    set_cell(ws, 3, 2, None, font=font_value, fill=fill_input)
    ws["B3"].number_format = "YYYY-MM-DD"
    set_cell(ws, 3, 3, "Value", font=font_white_b, fill=fill_navy)
    set_cell(ws, 3, 4, "Status", font=font_white_b, fill=fill_navy)
    # Also add a column header for A
    # Re-label C and D as column headers
    set_cell(ws, 3, 1, "Week of:", font=font_white_b, fill=fill_navy)
    ws.row_dimensions[3].height = 24

    # Freeze panes below row 3
    ws.freeze_panes = "A4"

    # ════════════════════════════════════════════════════════════════════
    #  SECTION 1 : Revenue & Prime Cost  (rows 4-10)
    # ════════════════════════════════════════════════════════════════════
    section_header(ws, 4, "1. Revenue & Prime Cost")
    label_row(ws, 5, "Weekly Revenue",               None, None, None, is_input=True, is_currency=True)
    label_row(ws, 6, "Food Cost $",                  None, None, None, is_input=True, is_currency=True)
    label_row(ws, 7, "Food Cost %",
              '=IF(B5=0,"",B6/B5)', "< 30%", None, is_pct=True)
    label_row(ws, 8, "Labor Cost $",                 None, None, None, is_input=True, is_currency=True)
    label_row(ws, 9, "Labor Cost %",
              '=IF(B5=0,"",B8/B5)', "< 30%", None, is_pct=True)
    label_row(ws, 10, "Prime Cost $",
              '=B6+B8', None, None, is_currency=True)
    # Prime Cost % with status in D
    label_row(ws, 11, "Prime Cost %",
              '=IF(B5=0,"",B10/B5)', "\u2264 60%",
              '=IF(B5=0,"",IF((B6+B8)/B5<=0.6,"PASS","OVER"))',
              is_pct=True)

    # ════════════════════════════════════════════════════════════════════
    #  SECTION 2 : Menu Performance  (rows 12-18)
    # ════════════════════════════════════════════════════════════════════
    section_header(ws, 12, "2. Menu Engineering")
    # Removed duplicate row 12 header
    label_row(ws, 13, "Stars (high profit, high pop.)", None, "Promote", None, is_input=True, is_integer=True)
    label_row(ws, 14, "Plowhorses (low profit, high pop.)", None, "Re-price", None, is_input=True, is_integer=True)
    label_row(ws, 15, "Puzzles (high profit, low pop.)", None, "Reposition", None, is_input=True, is_integer=True)
    label_row(ws, 16, "Dogs (low profit, low pop.)", None, "Cut / rework", None, is_input=True, is_integer=True)
    label_row(ws, 17, "Total Menu Items",
              '=B13+B14+B15+B16', None, None, is_integer=True)
    label_row(ws, 18, "Star % of Total",
              '=IF(B17=0,"",B13/B17)', "> 25%",
              '=IF(B17=0,"",IF(B13/B17>=0.25,"STRONG","LOW"))',
              is_pct=True)

    # ════════════════════════════════════════════════════════════════════
    #  SECTION 3 : Waste  (rows 20-25)  — skip row 19 as spacer
    # ════════════════════════════════════════════════════════════════════
    # Blank spacer row 19
    for c in range(1, 5):
        ws.cell(row=19, column=c).fill = fill_white
        ws.cell(row=19, column=c).border = Border()
    ws.row_dimensions[19].height = 6

    section_header(ws, 20, "3. Waste Tracking")
    label_row(ws, 21, "Weekly Waste $", None, None, None, is_input=True, is_currency=True)
    label_row(ws, 22, "Weekly Purchases $", None, None, None, is_input=True, is_currency=True)
    label_row(ws, 23, "Waste %",
              '=IF(B22=0,"",B21/B22)', "\u2264 4%", None, is_pct=True)
    label_row(ws, 24, "Waste Target", 0.04, None, None, is_pct=True)
    label_row(ws, 25, "Waste Status", None, None,
              '=IF(B22=0,"",IF(B21/B22<=0.04,"PASS","OVER"))')

    # ════════════════════════════════════════════════════════════════════
    #  SECTION 4 : Vendor Savings  (rows 27-31)
    # ════════════════════════════════════════════════════════════════════
    for c in range(1, 5):
        ws.cell(row=26, column=c).fill = fill_white
        ws.cell(row=26, column=c).border = Border()
    ws.row_dimensions[26].height = 6

    section_header(ws, 27, "4. Vendor Savings")
    label_row(ws, 28, "Current Monthly Vendor Spend", None, None, None, is_input=True, is_currency=True)
    label_row(ws, 29, "Best Available Price Total", None, None, None, is_input=True, is_currency=True)
    label_row(ws, 30, "Potential Monthly Savings",
              '=IF(B28=0,"",B28-B29)', None, None, is_currency=True)
    label_row(ws, 31, "Annual Savings Potential",
              '=B30*12', None, None, is_currency=True)

    # ════════════════════════════════════════════════════════════════════
    #  SECTION 5 : Delivery Profitability  (rows 33-37)
    # ════════════════════════════════════════════════════════════════════
    for c in range(1, 5):
        ws.cell(row=32, column=c).fill = fill_white
        ws.cell(row=32, column=c).border = Border()
    ws.row_dimensions[32].height = 6

    section_header(ws, 33, "5. Delivery Profitability")
    label_row(ws, 34, "Delivery Revenue", None, None, None, is_input=True, is_currency=True)
    label_row(ws, 35, "Delivery Costs (fees+pkg+labor)", None, None, None, is_input=True, is_currency=True)
    label_row(ws, 36, "Net Delivery Profit / Loss",
              '=B34-B35', None,
              '=IF(B34=0,"",IF(B34-B35>=0,"PROFIT","LOSS"))',
              is_currency=True)
    label_row(ws, 37, "Delivery Margin %",
              '=IF(B34=0,"",(B34-B35)/B34)', "> 15%", None, is_pct=True)

    # ════════════════════════════════════════════════════════════════════
    #  SECTION 6 : Inventory / Par  (rows 39-43)
    # ════════════════════════════════════════════════════════════════════
    for c in range(1, 5):
        ws.cell(row=38, column=c).fill = fill_white
        ws.cell(row=38, column=c).border = Border()
    ws.row_dimensions[38].height = 6

    section_header(ws, 39, "6. Inventory & Par Levels")
    label_row(ws, 40, "Items At Par", None, None, None, is_input=True, is_integer=True)
    label_row(ws, 41, "Items Below Par", None, None, None, is_input=True, is_integer=True)
    label_row(ws, 42, "Items Over Par", None, None, None, is_input=True, is_integer=True)
    label_row(ws, 43, "Par Compliance %",
              '=IF((B40+B41+B42)=0,"",B40/(B40+B41+B42))',
              "\u2265 85%",
              '=IF((B40+B41+B42)=0,"",IF(B40/(B40+B41+B42)>=0.85,"PASS","LOW"))',
              is_pct=True)

    # ════════════════════════════════════════════════════════════════════
    #  SECTION 7 : Menu Psychology Score  (rows 45-48)
    # ════════════════════════════════════════════════════════════════════
    for c in range(1, 5):
        ws.cell(row=44, column=c).fill = fill_white
        ws.cell(row=44, column=c).border = Border()
    ws.row_dimensions[44].height = 6

    section_header(ws, 45, "7. Menu Psychology Audit")
    label_row(ws, 46, "Audit Score (out of 100)", None, "\u2265 70", None, is_input=True, is_integer=True)
    label_row(ws, 47, "Score Rating",
              '=IF(B46="","",IF(B46>=90,"Optimized",IF(B46>=70,"Strong",IF(B46>=40,"Needs Work","Critical"))))',
              None, None)
    # Add data validation 0-100 on B46
    dv_score = DataValidation(type="whole", operator="between",
                              formula1=0, formula2=100,
                              errorTitle="Invalid Score",
                              error="Enter a whole number between 0 and 100.")
    dv_score.showErrorMessage = True
    ws.add_data_validation(dv_score)
    dv_score.add(ws["B46"])

    # Status colouring for score rating in D47
    label_row(ws, 48, "Score Benchmark",
              None, "70+ = Strong", None)

    # ════════════════════════════════════════════════════════════════════
    #  SECTION 8 : Weekly Health Score  (rows 50-55)
    # ════════════════════════════════════════════════════════════════════
    for c in range(1, 5):
        ws.cell(row=49, column=c).fill = fill_white
        ws.cell(row=49, column=c).border = Border()
    ws.row_dimensions[49].height = 6

    section_header(ws, 50, "8. Weekly Health Score (Composite)")

    # Sub-scores (auto-calc, 0-20 each, total 100)
    label_row(ws, 51, "Prime Cost Component (0-20)",
              '=IF(B5=0,0,IF(B10/B5<=0.6,20,IF(B10/B5<=0.65,10,0)))',
              "20 = \u2264 60%", None, is_integer=True)
    label_row(ws, 52, "Waste Component (0-20)",
              '=IF(B22=0,0,IF(B21/B22<=0.04,20,IF(B21/B22<=0.06,10,0)))',
              "20 = \u2264 4%", None, is_integer=True)
    label_row(ws, 53, "Par Compliance Component (0-20)",
              '=IF((B40+B41+B42)=0,0,IF(B40/(B40+B41+B42)>=0.85,20,IF(B40/(B40+B41+B42)>=0.7,10,0)))',
              "20 = \u2265 85%", None, is_integer=True)
    label_row(ws, 54, "Menu Psychology Component (0-20)",
              '=IF(B46="",0,IF(B46>=70,20,IF(B46>=40,10,0)))',
              "20 = score \u2265 70", None, is_integer=True)
    label_row(ws, 55, "Delivery Component (0-20)",
              '=IF(B34=0,0,IF((B34-B35)/B34>=0.15,20,IF((B34-B35)/B34>=0,10,0)))',
              "20 = margin \u2265 15%", None, is_integer=True)

    # Composite total
    label_row(ws, 56, "WEEKLY HEALTH SCORE",
              '=B51+B52+B53+B54+B55', "Target: 80+", None, is_integer=True)
    # Bold + larger font for the score cell
    ws["B56"].font = Font(name="Calibri", size=14, bold=True, color=NAVY)
    ws["A56"].font = Font(name="Calibri", size=12, bold=True, color=WHITE)

    label_row(ws, 57, "Health Rating",
              '=IF(B56>=90,"Optimized",IF(B56>=70,"Strong",IF(B56>=50,"Stable",IF(B56>=30,"At Risk","Critical"))))',
              None,
              '=IF(B56>=70,"HEALTHY",IF(B56>=50,"MONITOR","ACTION NEEDED"))')

    # ── Conditional formatting for status column (D) ───────────────────
    # PASS / STRONG / PROFIT / HEALTHY = green
    green_font = Font(bold=True, color=GREEN)
    red_font   = Font(bold=True, color=RED)
    green_fill_light = PatternFill(start_color="E6F4EA", end_color="E6F4EA", fill_type="solid")
    red_fill_light   = PatternFill(start_color="FDE8E8", end_color="FDE8E8", fill_type="solid")

    for status_cell in ["D11", "D25", "D43", "D57"]:
        ws.conditional_formatting.add(
            status_cell,
            CellIsRule(operator="equal", formula=['"PASS"'],
                       fill=green_fill_light, font=green_font))
        ws.conditional_formatting.add(
            status_cell,
            CellIsRule(operator="equal", formula=['"HEALTHY"'],
                       fill=green_fill_light, font=green_font))
        ws.conditional_formatting.add(
            status_cell,
            CellIsRule(operator="equal", formula=['"OVER"'],
                       fill=red_fill_light, font=red_font))
        ws.conditional_formatting.add(
            status_cell,
            CellIsRule(operator="equal", formula=['"LOW"'],
                       fill=red_fill_light, font=red_font))
        ws.conditional_formatting.add(
            status_cell,
            CellIsRule(operator="equal", formula=['"ACTION NEEDED"'],
                       fill=red_fill_light, font=red_font))

    for status_cell in ["D18"]:
        ws.conditional_formatting.add(
            status_cell,
            CellIsRule(operator="equal", formula=['"STRONG"'],
                       fill=green_fill_light, font=green_font))
        ws.conditional_formatting.add(
            status_cell,
            CellIsRule(operator="equal", formula=['"LOW"'],
                       fill=red_fill_light, font=red_font))

    for status_cell in ["D36"]:
        ws.conditional_formatting.add(
            status_cell,
            CellIsRule(operator="equal", formula=['"PROFIT"'],
                       fill=green_fill_light, font=green_font))
        ws.conditional_formatting.add(
            status_cell,
            CellIsRule(operator="equal", formula=['"LOSS"'],
                       fill=red_fill_light, font=red_font))

    # Conditional formatting for B47 (score rating text)
    ws.conditional_formatting.add(
        "B47",
        CellIsRule(operator="equal", formula=['"Optimized"'],
                   fill=green_fill_light, font=green_font))
    ws.conditional_formatting.add(
        "B47",
        CellIsRule(operator="equal", formula=['"Strong"'],
                   fill=green_fill_light, font=green_font))
    ws.conditional_formatting.add(
        "B47",
        CellIsRule(operator="equal", formula=['"Needs Work"'],
                   fill=PatternFill(start_color="FFF9E6", end_color="FFF9E6", fill_type="solid"),
                   font=Font(bold=True, color="B7791F")))
    ws.conditional_formatting.add(
        "B47",
        CellIsRule(operator="equal", formula=['"Critical"'],
                   fill=red_fill_light, font=red_font))

    # Conditional formatting for B57 (health rating)
    ws.conditional_formatting.add(
        "B57",
        CellIsRule(operator="equal", formula=['"Optimized"'],
                   fill=green_fill_light, font=green_font))
    ws.conditional_formatting.add(
        "B57",
        CellIsRule(operator="equal", formula=['"Strong"'],
                   fill=green_fill_light, font=green_font))
    ws.conditional_formatting.add(
        "B57",
        CellIsRule(operator="equal", formula=['"Stable"'],
                   fill=PatternFill(start_color="FFF9E6", end_color="FFF9E6", fill_type="solid"),
                   font=Font(bold=True, color="B7791F")))
    ws.conditional_formatting.add(
        "B57",
        CellIsRule(operator="equal", formula=['"At Risk"'],
                   fill=PatternFill(start_color="FDE8E8", end_color="FDE8E8", fill_type="solid"),
                   font=Font(bold=True, color="C53030")))
    ws.conditional_formatting.add(
        "B57",
        CellIsRule(operator="equal", formula=['"Critical"'],
                   fill=red_fill_light, font=red_font))

    # ── Print setup ────────────────────────────────────────────────────
    ws.print_area = "A1:D57"
    ws.page_setup.orientation = "portrait"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.print_title_rows = "1:3"

    # ── Footer ─────────────────────────────────────────────────────────
    for c in range(1, 5):
        ws.cell(row=58, column=c).fill = fill_white
        ws.cell(row=58, column=c).border = Border()
    ws.merge_cells("A59:D59")
    set_cell(ws, 59, 1,
             "thegrumpychef.ca  |  The Grumpy Chef Template Bundle  |  Yellow = Input  |  Gray = Auto-Calc",
             font=Font(name="Calibri", size=9, italic=True, color="888888"),
             fill=fill_white,
             alignment=Alignment(horizontal="center", vertical="center"),
             border=Border(top=Side(style="thin", color=GOLD)))
    for c in range(2, 5):
        ws.cell(row=59, column=c).fill = fill_white
        ws.cell(row=59, column=c).border = Border(top=Side(style="thin", color=GOLD))


# ═══════════════════════════════════════════════════════════════════════════
#  SHEET 2 : INSTRUCTIONS
# ═══════════════════════════════════════════════════════════════════════════
def build_instructions(wb):
    ws = wb.create_sheet("Instructions")

    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 65
    ws.column_dimensions["C"].width = 40

    # Title
    ws.merge_cells("A1:C1")
    set_cell(ws, 1, 1, "How to Use This Dashboard", font=font_title, fill=fill_navy_dark,
             alignment=Alignment(horizontal="left", vertical="center"),
             border=Border(bottom=Side(style="medium", color=GOLD)))
    for c in [2, 3]:
        ws.cell(row=1, column=c).fill = fill_navy_dark
        ws.cell(row=1, column=c).border = Border(bottom=Side(style="medium", color=GOLD))
    ws.row_dimensions[1].height = 36

    # Intro
    instructions = [
        ("", "QUICK START", ""),
        ("1.", "Open the Dashboard tab each week.", ""),
        ("2.", "Enter the date in the \"Week of\" cell (B3).", ""),
        ("3.", "Fill in every YELLOW cell with data from your templates.", ""),
        ("4.", "Gray cells auto-calculate. Do NOT overwrite them.", ""),
        ("5.", "Check the Status column (D) for instant pass/fail signals.", ""),
        ("6.", "Copy your weekly results to the Weekly Log tab to track trends.", ""),
        ("", "", ""),
        ("", "COLOR CODING", ""),
        ("", "Yellow background (#FFF2CC) = Input cell. You type here.", "Enter your data"),
        ("", "Gray background (#D9D9D9)  = Auto-calculated. Hands off.", "Formula-driven"),
        ("", "Green status = Passing target.", "On track"),
        ("", "Red status = Over target. Investigate.", "Needs attention"),
        ("", "", ""),
        ("", "TEMPLATE MAPPING", "Which template feeds this section"),
        ("", "Section 1: Revenue & Prime Cost", "01 Prime Cost Calculator"),
        ("", "Section 2: Menu Engineering", "02 Menu Engineering Matrix"),
        ("", "Section 3: Waste Tracking", "03 Waste Log & Tracker"),
        ("", "Section 4: Vendor Savings", "04 Vendor Comparison Tool"),
        ("", "Section 5: Delivery Profitability", "05 Delivery Profitability Calc"),
        ("", "Section 6: Inventory & Par Levels", "06 Inventory & Par Level Sheet"),
        ("", "Section 7: Menu Psychology Audit", "07 Menu Psychology Checklist"),
        ("", "Section 8: Weekly Health Score", "Auto-calculated from Sections 1-7"),
        ("", "", ""),
        ("", "THE 8 TEMPLATES IN THIS BUNDLE", "Purpose"),
        ("", "01 Prime Cost Calculator", "Track food + labor as % of revenue weekly"),
        ("", "02 Menu Engineering Matrix", "Classify every dish: Star, Plowhorse, Puzzle, Dog"),
        ("", "03 Waste Log & Tracker", "Log daily waste and calculate weekly waste %"),
        ("", "04 Vendor Comparison Tool", "Compare supplier pricing across your top items"),
        ("", "05 Delivery Profitability Calculator", "Know if delivery is making or losing money"),
        ("", "06 Inventory & Par Level Sheet", "Set par levels, track compliance"),
        ("", "07 Menu Psychology Checklist", "Score your menu design against 25+ proven tactics"),
        ("", "08 Restaurant Operations Dashboard", "THIS FILE \u2014 the connective tissue"),
        ("", "", ""),
        ("", "WEEKLY HEALTH SCORE BREAKDOWN", "Max 20 pts each = 100 total"),
        ("", "Prime Cost \u2264 60% = 20 pts, \u2264 65% = 10, >65% = 0", ""),
        ("", "Waste \u2264 4% = 20 pts, \u2264 6% = 10, >6% = 0", ""),
        ("", "Par Compliance \u2265 85% = 20, \u2265 70% = 10, <70% = 0", ""),
        ("", "Menu Psychology Score \u2265 70 = 20, \u2265 40 = 10, <40 = 0", ""),
        ("", "Delivery Margin \u2265 15% = 20, \u2265 0% = 10, <0% = 0", ""),
    ]

    for i, (col_a, col_b, col_c) in enumerate(instructions, start=3):
        row = i
        a_font = Font(name="Calibri", size=11, bold=True, color=NAVY) if col_a else None
        b_font = font_section if col_b in ("QUICK START", "COLOR CODING", "TEMPLATE MAPPING",
                                            "THE 8 TEMPLATES IN THIS BUNDLE",
                                            "WEEKLY HEALTH SCORE BREAKDOWN") else \
                 Font(name="Calibri", size=11, color=BLACK)
        b_fill = fill_navy if col_b in ("QUICK START", "COLOR CODING", "TEMPLATE MAPPING",
                                         "THE 8 TEMPLATES IN THIS BUNDLE",
                                         "WEEKLY HEALTH SCORE BREAKDOWN") else fill_white

        set_cell(ws, row, 1, col_a, font=a_font, fill=fill_white, border=None)
        set_cell(ws, row, 2, col_b, font=b_font, fill=b_fill, border=None,
                 alignment=align_left)
        set_cell(ws, row, 3, col_c,
                 font=Font(name="Calibri", size=10, italic=True, color="555555"),
                 fill=fill_white, border=None, alignment=align_left)

    # Print setup
    ws.page_setup.orientation = "portrait"
    ws.page_setup.fitToWidth = 1


# ═══════════════════════════════════════════════════════════════════════════
#  SHEET 3 : WEEKLY LOG
# ═══════════════════════════════════════════════════════════════════════════
def build_weekly_log(wb):
    ws = wb.create_sheet("Weekly Log")

    headers = [
        ("Week", 14),
        ("Revenue", 16),
        ("Food Cost %", 14),
        ("Labor Cost %", 14),
        ("Prime Cost %", 14),
        ("Waste %", 12),
        ("Par Compliance %", 16),
        ("Menu Psych Score", 16),
        ("Delivery Margin %", 16),
        ("Health Score", 14),
        ("Rating", 14),
    ]

    # Title row
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    set_cell(ws, 1, 1, "Weekly Performance Log \u2014 The Grumpy Chef",
             font=font_title, fill=fill_navy_dark,
             alignment=Alignment(horizontal="left", vertical="center"),
             border=Border(bottom=Side(style="medium", color=GOLD)))
    for c in range(2, len(headers) + 1):
        ws.cell(row=1, column=c).fill = fill_navy_dark
        ws.cell(row=1, column=c).border = Border(bottom=Side(style="medium", color=GOLD))
    ws.row_dimensions[1].height = 36

    # Subtitle
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(headers))
    set_cell(ws, 2, 1, "Track 12 weeks of data to spot trends. Copy numbers from Dashboard each week.",
             font=font_subtitle, fill=fill_white, alignment=align_left, border=None)
    for c in range(2, len(headers) + 1):
        ws.cell(row=2, column=c).fill = fill_white

    # Column headers (row 3)
    for col_idx, (hdr, width) in enumerate(headers, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width
        set_cell(ws, 3, col_idx, hdr, font=font_white_b, fill=fill_navy,
                 alignment=align_center)

    ws.freeze_panes = "A4"

    # 12 data rows (rows 4-15)
    for row in range(4, 16):
        week_num = row - 3
        set_cell(ws, row, 1, f"Week {week_num}",
                 font=Font(name="Calibri", size=11, bold=True, color=NAVY),
                 fill=fill_white, alignment=align_center)
        # Revenue (currency)
        set_cell(ws, row, 2, None, font=font_value, fill=fill_input,
                 alignment=align_center, num_fmt='$#,##0')
        # Pct columns (3-6, 7, 9)
        for pct_col in [3, 4, 5, 6, 7, 9]:
            set_cell(ws, row, pct_col, None, font=font_value, fill=fill_input,
                     alignment=align_center, num_fmt='0.0%')
        # Menu psych score (integer, col 8)
        set_cell(ws, row, 8, None, font=font_value, fill=fill_input,
                 alignment=align_center, num_fmt='#,##0')
        # Health score (integer, col 10)
        set_cell(ws, row, 10, None, font=font_value, fill=fill_input,
                 alignment=align_center, num_fmt='#,##0')
        # Rating (auto-calc from health score)
        set_cell(ws, row, 11,
                 f'=IF(J{row}="","",IF(J{row}>=90,"Optimized",IF(J{row}>=70,"Strong",IF(J{row}>=50,"Stable",IF(J{row}>=30,"At Risk","Critical")))))',
                 font=font_value_bold, fill=fill_calc, alignment=align_center)

    # Conditional formatting on Rating column (K4:K15)
    green_font_cf = Font(bold=True, color=GREEN)
    red_font_cf   = Font(bold=True, color=RED)
    amber_font_cf = Font(bold=True, color="B7791F")
    green_fill_cf = PatternFill(start_color="E6F4EA", end_color="E6F4EA", fill_type="solid")
    red_fill_cf   = PatternFill(start_color="FDE8E8", end_color="FDE8E8", fill_type="solid")
    amber_fill_cf = PatternFill(start_color="FFF9E6", end_color="FFF9E6", fill_type="solid")

    rating_range = "K4:K15"
    ws.conditional_formatting.add(
        rating_range,
        CellIsRule(operator="equal", formula=['"Optimized"'], fill=green_fill_cf, font=green_font_cf))
    ws.conditional_formatting.add(
        rating_range,
        CellIsRule(operator="equal", formula=['"Strong"'], fill=green_fill_cf, font=green_font_cf))
    ws.conditional_formatting.add(
        rating_range,
        CellIsRule(operator="equal", formula=['"Stable"'], fill=amber_fill_cf, font=amber_font_cf))
    ws.conditional_formatting.add(
        rating_range,
        CellIsRule(operator="equal", formula=['"At Risk"'], fill=red_fill_cf, font=red_font_cf))
    ws.conditional_formatting.add(
        rating_range,
        CellIsRule(operator="equal", formula=['"Critical"'], fill=red_fill_cf, font=red_font_cf))

    # Averages row
    avg_row = 17
    set_cell(ws, 16, 1, None, fill=fill_white, border=None)  # spacer
    for c in range(2, len(headers) + 1):
        ws.cell(row=16, column=c).fill = fill_white
        ws.cell(row=16, column=c).border = Border()

    set_cell(ws, avg_row, 1, "AVERAGES", font=font_white_b, fill=fill_navy, alignment=align_center)
    # Revenue avg
    set_cell(ws, avg_row, 2, '=IF(COUNTA(B4:B15)=0,"",AVERAGE(B4:B15))',
             font=font_value_bold, fill=fill_calc, alignment=align_center, num_fmt='$#,##0')
    # Pct avgs
    for pct_col in [3, 4, 5, 6, 7, 9]:
        col_letter = get_column_letter(pct_col)
        set_cell(ws, avg_row, pct_col,
                 f'=IF(COUNTA({col_letter}4:{col_letter}15)=0,"",AVERAGE({col_letter}4:{col_letter}15))',
                 font=font_value_bold, fill=fill_calc, alignment=align_center, num_fmt='0.0%')
    # Menu psych avg
    set_cell(ws, avg_row, 8,
             '=IF(COUNTA(H4:H15)=0,"",AVERAGE(H4:H15))',
             font=font_value_bold, fill=fill_calc, alignment=align_center, num_fmt='#,##0')
    # Health score avg
    set_cell(ws, avg_row, 10,
             '=IF(COUNTA(J4:J15)=0,"",AVERAGE(J4:J15))',
             font=font_value_bold, fill=fill_calc, alignment=align_center, num_fmt='#,##0')
    # Rating for average
    set_cell(ws, avg_row, 11,
             f'=IF(J{avg_row}="","",IF(J{avg_row}>=90,"Optimized",IF(J{avg_row}>=70,"Strong",IF(J{avg_row}>=50,"Stable",IF(J{avg_row}>=30,"At Risk","Critical")))))',
             font=font_value_bold, fill=fill_calc, alignment=align_center)

    # Print
    ws.print_area = f"A1:K{avg_row}"
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.print_title_rows = "1:3"


# ═══════════════════════════════════════════════════════════════════════════
#  MAIN
# ═══════════════════════════════════════════════════════════════════════════
def main():
    wb = Workbook()

    build_dashboard(wb)
    build_instructions(wb)
    build_weekly_log(wb)

    # Set Dashboard as the active sheet
    wb.active = 0

    output_path = r"C:\Users\cschi\Downloads\00_Restaurant_Operations_Dashboard.xlsx"
    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    wb.save(output_path)
    print(f"Dashboard saved to: {output_path}")
    print(f"  Sheets: {wb.sheetnames}")
    print(f"  File size: {os.path.getsize(output_path):,} bytes")


if __name__ == "__main__":
    main()
